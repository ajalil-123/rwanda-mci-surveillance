"""
app.py — Flask backend for Rwanda MCI Surveillance System
Works locally (Windows/Mac/Linux) and on Render.com
"""
import os, json, threading, time, logging
from datetime import datetime
from flask import Flask, jsonify, send_from_directory, request

# ── paths ─────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# On Render, use /tmp for writable storage (persists within a session).
# For permanent persistence, mount a Render Disk at /data.
# Locally, use the project folder.
if os.path.exists("/tmp") and os.environ.get("RENDER"):
    DATA_DIR = "/data" if os.path.exists("/data") else "/tmp"
    LOG_DIR  = "/tmp/logs"
else:
    DATA_DIR = os.path.join(BASE_DIR, "data")
    LOG_DIR  = os.path.join(BASE_DIR, "logs")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(LOG_DIR,  exist_ok=True)

# Inject paths into database module before importing
os.environ.setdefault("MCI_DATA_DIR", DATA_DIR)
os.environ.setdefault("MCI_LOG_DIR",  LOG_DIR)

# ── logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(LOG_DIR, "system.log"), encoding="utf-8"),
        logging.StreamHandler(),
    ]
)
logger = logging.getLogger("app")

# ── app ───────────────────────────────────────────────────────────────────────
# Use the directory containing this file — works correctly under gunicorn on Render
APP_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, static_folder=APP_DIR, template_folder=APP_DIR)

try:
    from flask_cors import CORS; CORS(app)
except ImportError:
    pass

# ── local modules ─────────────────────────────────────────────────────────────
from database import init_db, get_db
from scraper  import run_historical_scrape, run_incremental_scrape
from analytics import (
    get_summary_stats, by_incident_type, district_hotspots,
    monthly_trend, yearly_trend, seasonal_pattern, type_trend_by_year,
    predict_next_month, high_risk_districts,
    recent_incidents, all_mapped_incidents, scrape_status,
    monthly_heatmap, day_of_week_pattern, case_fatality_rate,
    deadliest_incidents, year_over_year, province_trend,
    hour_of_day_pattern, peak_months_summary,
    by_source_tier, sources_by_tier,
)

# ── background scheduler ──────────────────────────────────────────────────────
_scheduler_running = False
_historical_done   = False
_last_incremental  = None
_scrape_lock       = threading.Lock()

def scheduler():
    global _scheduler_running, _historical_done, _last_incremental
    _scheduler_running = True
    logger.info("Scheduler started.")

    # ── Step 1: historical scrape (runs once per DB) ──────────────────────
    conn = get_db()
    count = conn.execute("SELECT COUNT(*) FROM incidents").fetchone()[0]
    conn.close()

    if count == 0:
        logger.info("Empty database — running historical scrape first.")
        with _scrape_lock:
            run_historical_scrape()
        _historical_done = True
    else:
        logger.info(f"Database has {count} incidents — skipping historical scrape.")
        _historical_done = True

    # ── Step 2: incremental refresh every 30 minutes ─────────────────────
    while _scheduler_running:
        logger.info("Running incremental refresh...")
        with _scrape_lock:
            run_incremental_scrape()
        _last_incremental = datetime.utcnow().isoformat()
        logger.info("Incremental refresh done. Next run in 30 minutes.")
        time.sleep(1800)

# ── API routes ────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(APP_DIR, "index.html")

# Dashboard stats
@app.route("/api/stats")
def api_stats():
    return jsonify(get_summary_stats())

# All incidents for map
@app.route("/api/incidents/map")
def api_map():
    min_deaths = int(request.args.get("min_deaths", 0))
    return jsonify(all_mapped_incidents(min_deaths))

# Recent incident feed
@app.route("/api/incidents/recent")
def api_recent():
    hours      = int(request.args.get("hours", 72))
    min_sev    = int(request.args.get("min_severity", 1))
    return jsonify(recent_incidents(hours, min_sev))

# Single incident
@app.route("/api/incidents/<int:iid>")
def api_incident(iid):
    conn = get_db()
    row  = conn.execute("SELECT * FROM incidents WHERE id=?", (iid,)).fetchone()
    conn.close()
    return jsonify(dict(row)) if row else (jsonify({"error": "Not found"}), 404)

# Resolve / reopen
@app.route("/api/incidents/<int:iid>/status", methods=["POST"])
def api_set_status(iid):
    status = request.json.get("status","resolved")
    conn   = get_db()
    conn.execute("UPDATE incidents SET status=? WHERE id=?", (status, iid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

# Analytics
@app.route("/api/analytics/types")
def api_types():
    return jsonify(by_incident_type())

@app.route("/api/analytics/hotspots")
def api_hotspots():
    return jsonify(district_hotspots())

@app.route("/api/analytics/monthly")
def api_monthly():
    years = int(request.args.get("years", 5))
    return jsonify(monthly_trend(years))

@app.route("/api/analytics/yearly")
def api_yearly():
    return jsonify(yearly_trend())

@app.route("/api/analytics/seasonal")
def api_seasonal():
    return jsonify(seasonal_pattern())

@app.route("/api/analytics/type-trend")
def api_type_trend():
    return jsonify(type_trend_by_year())

@app.route("/api/analytics/heatmap")
def api_heatmap():
    return jsonify(monthly_heatmap())

@app.route("/api/analytics/dow")
def api_dow():
    return jsonify(day_of_week_pattern())

@app.route("/api/analytics/cfr")
def api_cfr():
    return jsonify(case_fatality_rate())

@app.route("/api/analytics/deadliest")
def api_deadliest():
    limit = int(request.args.get("limit", 20))
    return jsonify(deadliest_incidents(limit))

@app.route("/api/analytics/yoy")
def api_yoy():
    return jsonify(year_over_year())

@app.route("/api/analytics/province-trend")
def api_province_trend():
    return jsonify(province_trend())

@app.route("/api/analytics/peak-months")
def api_peak_months():
    return jsonify(peak_months_summary())

@app.route("/api/analytics/source-tiers")
def api_source_tiers():
    return jsonify(by_source_tier())

@app.route("/api/analytics/sources-by-tier")
def api_sources_by_tier():
    return jsonify(sources_by_tier())

# Predictions
@app.route("/api/predictions/next-month")
def api_predict_month():
    return jsonify(predict_next_month())

@app.route("/api/predictions/risk-districts")
def api_risk_districts():
    return jsonify(high_risk_districts())

# Scraper control
@app.route("/api/scraper/status")
def api_scraper_status():
    conn  = get_db()
    total = conn.execute("SELECT COUNT(*) FROM incidents").fetchone()[0]
    hist  = conn.execute("SELECT COUNT(*) FROM incidents WHERE is_historical=1").fetchone()[0]
    conn.close()
    return jsonify({
        "scheduler_running":   _scheduler_running,
        "historical_complete": _historical_done,
        "last_incremental":    _last_incremental,
        "total_incidents":     total,
        "historical_incidents":hist,
        "scrape_log":          scrape_status()[:10],
    })

@app.route("/api/scraper/refresh", methods=["POST"])
def api_manual_refresh():
    """Manually trigger an incremental refresh."""
    if _scrape_lock.locked():
        return jsonify({"ok": False, "message": "Scrape already running"}), 429
    def do():
        with _scrape_lock:
            run_incremental_scrape()
    threading.Thread(target=do, daemon=True).start()
    return jsonify({"ok": True, "message": "Incremental refresh started"})

@app.route("/api/scraper/historical", methods=["POST"])
def api_manual_historical():
    """Re-run full historical scrape (admin use)."""
    if _scrape_lock.locked():
        return jsonify({"ok": False, "message": "Scrape already running"}), 429
    def do():
        with _scrape_lock:
            run_historical_scrape()
    threading.Thread(target=do, daemon=True).start()
    return jsonify({"ok": True, "message": "Historical scrape started"})

# ── Data Explorer ─────────────────────────────────────────────────────────────
@app.route("/api/data/all")
def api_data_all():
    """Return all incidents for the Data Explorer table."""
    conn = get_db()
    rows = conn.execute("""
        SELECT id, event_date, detected_at, title, incident_type,
               district, province, latitude, longitude,
               deaths, injured, missing, severity,
               source_name, source_url, source_tier, media_type,
               ai_summary, ai_confidence, status, is_historical
        FROM incidents
        ORDER BY COALESCE(event_date, detected_at) DESC
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ── Exports ───────────────────────────────────────────────────────────────────
@app.route("/api/export/csv")
def api_export_csv():
    """Download all incidents as CSV (filtered by query params)."""
    import csv, io
    rows = _filtered_rows(request.args)
    si   = io.StringIO()
    cols = ["id","event_date","title","incident_type","district","province",
            "deaths","injured","missing","severity","source_name","source_tier",
            "source_url","media_type","detected_at","ai_summary"]
    writer = csv.DictWriter(si, fieldnames=cols, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    output = si.getvalue()
    from flask import Response
    return Response(
        "\ufeff" + output,          # BOM for Excel UTF-8
        mimetype="text/csv",
        headers={"Content-Disposition":
                 f"attachment; filename=rwanda_mci_{datetime.now().strftime('%Y%m%d')}.csv"}
    )

@app.route("/api/export/excel")
def api_export_excel():
    """Download incidents as Excel .xlsx file."""
    import io
    rows = _filtered_rows(request.args)
    from flask import Response

    # Build Excel manually using openpyxl if available, else fall back to CSV
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rwanda MCI Incidents"

        cols = ["id","event_date","title","incident_type","district","province",
                "deaths","injured","missing","severity","source_name","source_tier",
                "source_url","media_type","detected_at","ai_summary"]
        headers = ["ID","Date","Title","Type","District","Province",
                   "Deaths","Injured","Missing","Severity","Source","Tier",
                   "URL","Media","Detected At","AI Summary"]

        # header row styling
        header_fill = PatternFill("solid", fgColor="0D1117")
        for ci, (col, hdr) in enumerate(zip(cols, headers), 1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font      = Font(bold=True, color="58A6FF", name="Calibri", size=10)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        # severity colours
        sev_colours = {1:"3FB950",2:"D29922",3:"DB6D28",4:"F85149",5:"BC8CFF"}

        for ri, row in enumerate(rows, 2):
            for ci, col in enumerate(cols, 1):
                val  = row.get(col, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = Font(name="Calibri", size=9)
                if col == "deaths" and (val or 0) > 0:
                    cell.font = Font(name="Calibri", size=9, bold=True, color="F85149")
                if col == "severity":
                    sev_col = sev_colours.get(val, "484F58")
                    cell.fill = PatternFill("solid", fgColor=sev_col)

        # column widths
        widths = [5,12,55,16,14,12,7,7,7,5,22,5,40,12,18,60]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":
                     f"attachment; filename=rwanda_mci_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )

    except ImportError:
        # openpyxl not installed — redirect to CSV
        return api_export_csv()

def _filtered_rows(args):
    """Apply Data Explorer filters and return list of dicts."""
    q      = args.get("q","").lower()
    type_  = args.get("type","")
    prov   = args.get("province","")
    year   = args.get("year","")
    mind   = int(args.get("min_deaths",0) or 0)

    conn   = get_db()
    rows   = conn.execute("""
        SELECT id, event_date, detected_at, title, incident_type,
               district, province, deaths, injured, missing, severity,
               source_name, source_url, media_type, ai_summary
        FROM incidents
        ORDER BY COALESCE(event_date, detected_at) DESC
    """).fetchall()
    conn.close()

    result = []
    for r in rows:
        d = dict(r)
        text = f"{d.get('title','')} {d.get('district','')} {d.get('province','')} {d.get('incident_type','')} {d.get('source_name','')}".lower()
        dy   = (d.get("event_date") or d.get("detected_at") or "")[:4]
        if q    and q    not in text:       continue
        if type_ and d.get("incident_type") != type_: continue
        if prov  and d.get("province")      != prov:  continue
        if year  and dy                     != year:  continue
        if (d.get("deaths") or 0) < mind:             continue
        result.append(d)
    return result

# ── module-level initialisation (runs under gunicorn too) ─────────────────────
# Create database schema and start background scheduler when the module is
# imported. This ensures gunicorn workers have a working DB before serving
# any requests. Guarded with a flag so it only runs once even with --preload.
_INITIALISED = False

def _initialise():
    global _INITIALISED
    if _INITIALISED:
        return
    init_db()
    t = threading.Thread(target=scheduler, daemon=True)
    t.start()
    _INITIALISED = True
    logger.info("Database initialised and scheduler started.")

_initialise()

# ── entry point (only used when running `python app.py` directly) ─────────────
if __name__ == "__main__":
    logger.info("Server starting on http://localhost:5050")
    app.run(host="0.0.0.0", port=5050, debug=False, use_reloader=False)